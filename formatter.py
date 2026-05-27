"""
formatter.py — Writes extracted rows to a formatted .xlsx file.
"""

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_BG   = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_ROW_BG  = PatternFill("solid", fgColor="D9E2F3")  # light blue alternating rows
_MAX_COL_WIDTH = 50


class ExcelFormatter:
    """
    Writes a list of row-lists to a single Excel worksheet.

    Features:
    - Bold blue header row with white text
    - Alternating row shading for readability
    - Auto-sized column widths (capped at 50 chars)
    - Frozen header row (row 1 stays visible while scrolling)
    - Auto-filter on header row
    """

    def __init__(
        self,
        sheet_name: str = "Data",
        logger: logging.Logger | None = None,
    ) -> None:
        self.sheet_name = sheet_name
        self.log = logger or logging.getLogger(__name__)
        self._wb = openpyxl.Workbook()
        self._ws = self._wb.active
        self._ws.title = self.sheet_name

    # ── public ───────────────────────────────────────────────────────────────

    def write(self, rows: list[list[str]]) -> None:
        """Append all rows then apply formatting."""
        if not rows:
            self.log.warning("No rows to write.")
            return

        for idx, row in enumerate(rows, start=1):
            self._ws.append(row)
            if idx == 1:
                self._format_header(len(row))
            else:
                if idx % 2 == 0:
                    self._shade_row(idx, len(row))
                self._apply_wrap(idx, len(row))

        self._auto_column_widths()
        self._ws.freeze_panes = "A2"
        self._ws.auto_filter.ref = self._ws.dimensions
        self.log.info("Formatted %d rows (%d columns).", len(rows), len(rows[0]))

    def save(self, output_path: str | Path) -> None:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(path)
        self.log.info("Saved: %s", path)

    # ── private ──────────────────────────────────────────────────────────────

    def _format_header(self, num_cols: int) -> None:
        for col in range(1, num_cols + 1):
            cell = self._ws.cell(row=1, column=col)
            cell.fill = _HEADER_BG
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self._ws.row_dimensions[1].height = 30

    def _shade_row(self, row_idx: int, num_cols: int) -> None:
        for col in range(1, num_cols + 1):
            self._ws.cell(row=row_idx, column=col).fill = _ALT_ROW_BG

    def _apply_wrap(self, row_idx: int, num_cols: int) -> None:
        """Set wrap_text=True and vertical=top for cells that contain newlines (Alt+Enter)."""
        for col in range(1, num_cols + 1):
            cell = self._ws.cell(row=row_idx, column=col)
            if cell.value and "\n" in str(cell.value):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _auto_column_widths(self) -> None:
        for col in self._ws.columns:
            max_len = max(
                (
                    # For multi-line cells use the longest single line, not total length
                    max((len(ln) for ln in str(cell.value).splitlines()), default=0)
                    for cell in col
                    if cell.value is not None
                ),
                default=8,
            )
            letter = get_column_letter(col[0].column)
            self._ws.column_dimensions[letter].width = min(max_len + 4, _MAX_COL_WIDTH)
