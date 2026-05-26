"""
tests/test_formatter.py
Unit tests for ExcelFormatter — no PDF required.

Run:  pytest tests/test_formatter.py -v
"""

import sys
import tempfile
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from formatter import ExcelFormatter


SAMPLE_ROWS = [
    ["ลำดับ", "ชื่อ",       "แผนก",    "เงินเดือน"],
    ["1",     "สมชาย ใจดี", "บัญชี",   "25000"],
    ["2",     "สมหญิง รัก", "บุคคล",   "28000"],
    ["3",     "วิชัย เก่ง",  "IT",      "35000"],
]


class TestExcelFormatterInit:

    def test_creates_workbook(self):
        fmt = ExcelFormatter()
        assert fmt._wb is not None

    def test_creates_worksheet(self):
        fmt = ExcelFormatter()
        assert fmt._ws is not None

    def test_default_sheet_name(self):
        fmt = ExcelFormatter()
        assert fmt._ws.title == "Data"

    def test_custom_sheet_name(self):
        fmt = ExcelFormatter(sheet_name="ยอดขาย")
        assert fmt._ws.title == "ยอดขาย"


class TestExcelFormatterWrite:

    def test_row_count(self):
        """All rows are written to the sheet."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        assert fmt._ws.max_row == len(SAMPLE_ROWS)

    def test_column_count(self):
        """Column count matches the widest row."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        assert fmt._ws.max_column == len(SAMPLE_ROWS[0])

    def test_header_cell_values(self):
        """First row values match input header."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        header = [fmt._ws.cell(1, c).value for c in range(1, 5)]
        assert header == SAMPLE_ROWS[0]

    def test_data_cell_values(self):
        """Data rows are written correctly."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        assert fmt._ws.cell(2, 2).value == "สมชาย ใจดี"
        assert fmt._ws.cell(3, 3).value == "บุคคล"
        assert fmt._ws.cell(4, 4).value == "35000"

    def test_freeze_pane(self):
        """Freeze pane is set at A2 (header stays visible)."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        assert str(fmt._ws.freeze_panes) == "A2"

    def test_auto_filter_set(self):
        """Auto-filter is enabled."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        assert fmt._ws.auto_filter.ref is not None

    def test_header_has_fill(self):
        """Header row cells have a background fill colour."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        fill = fmt._ws.cell(1, 1).fill
        assert fill.patternType == "solid"
        assert fill.fgColor.rgb[-6:] == "2E75B6"   # blue accent colour (ignore alpha prefix)

    def test_header_font_bold_white(self):
        """Header font is bold and white."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        font = fmt._ws.cell(1, 1).font
        assert font.bold is True
        assert font.color.rgb[-6:] == "FFFFFF"   # white (ignore alpha prefix)

    def test_column_widths_set(self):
        """Column widths are auto-sized (not left at default 0)."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        from openpyxl.utils import get_column_letter
        width = fmt._ws.column_dimensions[get_column_letter(1)].width
        assert width > 0

    def test_column_width_capped_at_50(self):
        """Column width never exceeds 50 characters."""
        long_rows = [["x" * 200], ["y" * 200]]
        fmt = ExcelFormatter()
        fmt.write(long_rows)
        from openpyxl.utils import get_column_letter
        width = fmt._ws.column_dimensions[get_column_letter(1)].width
        assert width <= 50

    def test_empty_rows_logs_warning(self, caplog):
        """write() with no rows logs a warning and doesn't crash."""
        import logging
        fmt = ExcelFormatter()
        with caplog.at_level(logging.WARNING):
            fmt.write([])
        assert fmt._ws.max_row <= 1   # nothing written


class TestExcelFormatterSave:

    def test_save_creates_file(self):
        """save() writes a real .xlsx file to disk."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = Path(f.name)
        try:
            fmt.save(out)
            assert out.exists()
            assert out.stat().st_size > 0
        finally:
            out.unlink(missing_ok=True)

    def test_saved_file_is_valid_xlsx(self):
        """File saved by formatter can be re-opened by openpyxl."""
        import openpyxl
        fmt = ExcelFormatter(sheet_name="TestSheet")
        fmt.write(SAMPLE_ROWS)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = Path(f.name)
        try:
            fmt.save(out)
            wb = openpyxl.load_workbook(out)
            assert "TestSheet" in wb.sheetnames
            ws = wb["TestSheet"]
            assert ws.cell(1, 1).value == "ลำดับ"
            assert ws.max_row == len(SAMPLE_ROWS)
        finally:
            out.unlink(missing_ok=True)

    def test_save_creates_parent_dirs(self):
        """save() creates missing parent directories automatically."""
        fmt = ExcelFormatter()
        fmt.write(SAMPLE_ROWS)
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "subdir" / "output.xlsx"
            fmt.save(out)
            assert out.exists()
